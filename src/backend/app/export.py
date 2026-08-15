from __future__ import annotations

import io
from xml.sax.saxutils import escape

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer

from .models import Report

# Mirrors ReportCard.tsx section-for-section, so a printed/exported report reads as
# the same document a student saw on screen, not a different summary of it.


def render_pdf(report: Report, content_title: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        title=f"Sockrates' FOOT-Notes - {content_title}",
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )
    styles = getSampleStyleSheet()
    body = styles["BodyText"]
    quote_style = ParagraphStyle(
        "Quote",
        parent=body,
        leftIndent=14,
        textColor=colors.HexColor("#5b6b7c"),
        fontName="Helvetica-Oblique",
    )

    # Everything below is LLM output or verbatim student text (evidence_quote
    # especially), so it must be XML-escaped before it reaches Paragraph markup -
    # an unescaped "<" or "&" either mis-renders or breaks the mini-XML parser.
    story = [
        Paragraph("Sockrates' FOOT-Notes", styles["Title"]),
        Paragraph(escape(content_title), styles["Heading2"]),
        Spacer(1, 6),
        Paragraph(f"Overall score: {report.overall_score} / 10", styles["Heading3"]),
        Paragraph(f"Bloom level reached: {escape(report.bloom_level_reached)}", body),
        Spacer(1, 10),
        Paragraph(escape(report.explanation), body),
        Spacer(1, 12),
    ]

    if report.accommodations:
        story.append(
            Paragraph(
                "Accommodations in effect: "
                f"{escape(', '.join(report.accommodations))}. These change how the content was "
                "presented, not what was asked of the student. Spelling and message "
                "length are never scored.",
                body,
            )
        )
        story.append(Spacer(1, 12))

    story.append(Paragraph("Rubric breakdown", styles["Heading2"]))
    for dim in report.dimensions:
        story.append(Paragraph(f"{escape(dim.name)} — {dim.score}/4", styles["Heading4"]))
        story.append(Paragraph(f"“{escape(dim.evidence_quote)}”", quote_style))
        story.append(Paragraph(escape(dim.feedback), body))
        story.append(Spacer(1, 8))

    story.append(Paragraph("Try this next time", styles["Heading2"]))
    story.append(Paragraph(escape(report.next_step), body))

    if not report.generated_with_llm:
        story.append(Spacer(1, 12))
        story.append(Paragraph("Generated in offline scoring mode.", styles["Italic"]))

    doc.build(story)
    return buffer.getvalue()


def render_xlsx(report: Report, content_title: str) -> bytes:
    wb = Workbook()
    bold = Font(bold=True)
    wrapped = Alignment(wrap_text=True, vertical="top")

    summary = wb.active
    summary.title = "Summary"
    rows = [
        ("Content", content_title),
        ("Session ID", report.session_id),
        ("Overall score", f"{report.overall_score} / 10"),
        ("Bloom level reached", report.bloom_level_reached),
        ("Explanation", report.explanation),
        ("Next step", report.next_step),
        ("Generated with LLM", "Yes" if report.generated_with_llm else "No"),
        ("Accommodations", ", ".join(report.accommodations) or "None"),
    ]
    for row_index, (label, value) in enumerate(rows, start=1):
        summary.cell(row=row_index, column=1, value=label).font = bold
        summary.cell(row=row_index, column=2, value=value).alignment = wrapped
    summary.column_dimensions["A"].width = 20
    summary.column_dimensions["B"].width = 70

    dims = wb.create_sheet("Dimensions")
    for col_index, header in enumerate(["Dimension", "Score", "Evidence Quote", "Feedback"], start=1):
        dims.cell(row=1, column=col_index, value=header).font = bold
    for row_index, dim in enumerate(report.dimensions, start=2):
        dims.cell(row=row_index, column=1, value=dim.name)
        dims.cell(row=row_index, column=2, value=f"{dim.score}/4")
        dims.cell(row=row_index, column=3, value=dim.evidence_quote).alignment = wrapped
        dims.cell(row=row_index, column=4, value=dim.feedback).alignment = wrapped
    for col_index, width in enumerate([22, 8, 45, 45], start=1):
        dims.column_dimensions[get_column_letter(col_index)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
