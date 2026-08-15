from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pypdf import PdfReader

from app import config, export
from app.main import app
from app.models import DimensionScore, Report

client = TestClient(app)


@pytest.fixture(autouse=True)
def offline_mode(monkeypatch):
    """Force stub mode so tests never depend on network access or a real token."""
    monkeypatch.setattr(config, "GITHUB_TOKEN", "")


def completed_session() -> str:
    sid = client.post("/session", json={}).json()["session_id"]
    reply = None
    for _ in range(config.MAX_TURNS):
        reply = client.post(
            "/chat",
            json={"session_id": sid, "message": "Why does this happen and what else could explain it?"},
        ).json()
    assert reply["is_complete"] is True
    return sid


TRICKY_REPORT = Report(
    session_id="abc123",
    overall_score=8,
    bloom_level_reached="Analyze",
    explanation="You moved from asking what happened to asking why.",
    dimensions=[
        DimensionScore(
            dimension="question_quality",
            name="Question Quality",
            score=3,
            evidence_quote="Why <would> they only ask 10 & not more?",
            feedback="Good probing question.",
        ),
    ],
    next_step="Try naming the assumption next time.",
    generated_with_llm=False,
    accommodations=["Dyslexia-friendly reading mode"],
)


# --- export.py: renderer correctness, including XML-unsafe student text -----------


def test_render_pdf_escapes_xml_special_characters_from_student_text():
    pdf_bytes = export.render_pdf(TRICKY_REPORT, "Do Phones Hurt Test Scores?")
    assert pdf_bytes.startswith(b"%PDF")
    reader = PdfReader(io.BytesIO(pdf_bytes))
    text = "".join(page.extract_text() for page in reader.pages)
    assert "Why <would> they only ask 10 & not more?" in text
    assert "8 / 10" in text
    assert "Analyze" in text
    assert "Dyslexia-friendly reading mode" in text


def test_render_xlsx_contains_summary_and_dimensions():
    xlsx_bytes = export.render_xlsx(TRICKY_REPORT, "Do Phones Hurt Test Scores?")
    assert xlsx_bytes.startswith(b"PK")
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    assert wb.sheetnames == ["Summary", "Dimensions"]

    summary_rows = dict(wb["Summary"].iter_rows(values_only=True))
    assert summary_rows["Overall score"] == "8 / 10"
    assert summary_rows["Bloom level reached"] == "Analyze"

    dim_rows = list(wb["Dimensions"].iter_rows(values_only=True))
    assert dim_rows[0] == ("Dimension", "Score", "Evidence Quote", "Feedback")
    assert dim_rows[1] == (
        "Question Quality",
        "3/4",
        "Why <would> they only ask 10 & not more?",
        "Good probing question.",
    )


# --- endpoints ----------------------------------------------------------------------


def test_report_pdf_unknown_session_returns_404():
    resp = client.get("/report/does-not-exist/pdf")
    assert resp.status_code == 404


def test_report_xlsx_unknown_session_returns_404():
    resp = client.get("/report/does-not-exist/xlsx")
    assert resp.status_code == 404


def test_report_pdf_before_completion_returns_409():
    sid = client.post("/session", json={}).json()["session_id"]
    resp = client.get(f"/report/{sid}/pdf")
    assert resp.status_code == 409


def test_report_xlsx_before_completion_returns_409():
    sid = client.post("/session", json={}).json()["session_id"]
    resp = client.get(f"/report/{sid}/xlsx")
    assert resp.status_code == 409


def test_report_pdf_endpoint_returns_downloadable_pdf_matching_the_report():
    sid = completed_session()
    report = client.get(f"/report/{sid}").json()

    resp = client.get(f"/report/{sid}/pdf")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert f"sockrates-report-{sid}.pdf" in resp.headers["content-disposition"]

    reader = PdfReader(io.BytesIO(resp.content))
    text = "".join(page.extract_text() for page in reader.pages)
    assert f"{report['overall_score']} / 10" in text
    assert report["bloom_level_reached"] in text
    for dim in report["dimensions"]:
        assert dim["evidence_quote"] in text


def test_report_xlsx_endpoint_returns_downloadable_xlsx_matching_the_report():
    sid = completed_session()
    report = client.get(f"/report/{sid}").json()

    resp = client.get(f"/report/{sid}/xlsx")
    assert resp.status_code == 200
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert f"sockrates-report-{sid}.xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.content))
    summary_rows = dict(wb["Summary"].iter_rows(values_only=True))
    assert summary_rows["Overall score"] == f"{report['overall_score']} / 10"
    assert summary_rows["Bloom level reached"] == report["bloom_level_reached"]

    dim_rows = list(wb["Dimensions"].iter_rows(values_only=True))[1:]
    exported_quotes = {row[2] for row in dim_rows}
    assert exported_quotes == {dim["evidence_quote"] for dim in report["dimensions"]}
